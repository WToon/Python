from openpyxl import load_workbook
import pickle

# Initialization

workbook = "DataMine.xlsx"
dataTypes = ["Age", "Grade", "Class", "Gender"]

wb = load_workbook(workbook, data_only=True)
ws = wb.active

# data reading methods


def get_data_headers():
    data_headers = {}
    for row in range(1, 10):
        for column in range(1, 20):
            cell_value = ws.cell(row=row, column=column).value
            for head in dataTypes:
                if head == cell_value:
                    data_headers[head] = str(row)+str(column)
    return data_headers


def collect_data():
    all_data = {}
    headers = get_data_headers()
    for head in headers.keys():
        data = []
        header = headers.get(head)
        s_row = header[0]
        s_column = header[1]
        for i in range(int(s_row)+1, 1000):
            val = ws.cell(row=i, column=int(s_column)).value
            if val is None:
                break
            else:
                data.append(val)
        all_data[str(head)] = data
    return all_data

# Storing and loading data as pickles


def pickle_this_sheet():
    pickle.dump(collect_data(), open("PickledData\save2.pickle", "wb"))


def open_this_pickle(name):
    name = 'PickledData\save.pickle'
    with open(name, 'rb') as f:
        data = pickle.load(f)
    return data

print(collect_data())

